#!/usr/bin/python3
# -*- coding: utf-8 -*-
# encoding=utf8
# 20160623 compare DPI detail XDR to PGW rule
# support ZTE EPC DPI FILE
# inputfile should be  Excel 2010 files (ie: .xlsx) and not filter used, older Excel files (ie: .xls) were not supported
# yao.savage@gmail.com
# (DONE) 20160628 IP match
# (DONE) 20160701 URL match
# (DONE) 20170930 dpi file format change to new version ,RuleBaseID  column from field 27 ro 30
# (DONE) 20190315 dpi file format change to new version ,RuleBaseID  column from field 30 ro 31
# (DONE) 20200312 change from Python27 to Python37
# (DONE) 20200312 IPv6 support add(Use the ipset.size instead of len of list)
# (TODO) CASE 20200331  18156461531 RG 3532706100
import netaddr
import re
from openpyxl import load_workbook
from openpyxl.workbook import workbook
from datetime import datetime
import os
import warnings

warnings.filterwarnings("ignore")


def genrepatten(inurllist):
    pattenlist = []
    if len(inurllist) > 0:
        for everyurl in inurllist:
            # patten1 = everyurl.strip().replace('.', '\.')
            patten1 = everyurl.strip().replace('.', '\\.')
            if everyurl.startswith("*"):
                flag = 1
            else:
                flag = 0
            patten2 = patten1.replace('*', '.*')
            if flag == 1:
                # patten2 = u'://[^\/]' + patten2[1::]
                patten2 = u'://[^\\/]' + patten2[1::]
            pattenlist.append(patten2)
    return pattenlist


def definesubnet(insubnetlist):
    fnsubnet = []
    if len(insubnetlist) > 0:
        for everysubnet in insubnetlist:
            mynet = netaddr.IPNetwork(everysubnet)
            fnsubnet.append(mynet)
        mergesubnet = netaddr.cidr_merge(list(set(fnsubnet)))
        return mergesubnet
    else:
        return fnsubnet


def copyrow(inws, outws, inrow, outrow, columnnum):
    for everycolumn in range(1, columnnum + 1):
        outws.cell(row=outrow, column=everycolumn).value = inws.cell(row=inrow, column=everycolumn).value


def genipandhost(inip, inhost):
    if len(inip) > 0:
        print(u'未匹配上的IP地址有:', end=' ')
        for everyip in inip:
            print(everyip, end=',')
    print('\r')
    if len(inhost) > 0:
        print(u'未匹配上的主机名有:', end=' ')
        for everyhost in inhost:
            print(everyhost, end=',')


def exportrule(outlist, outfilename):
    if len(outlist) > 0:
        outfile = open(outfilename, "w")
        # outfile = open(outfilename.decode('utf-8'), "w")
        for everyele in outlist:
            print(everyele, file=outfile)
        outfile.close()


# ZTE xGW DPI file
dpifile = 'DPI-20200404.xlsx'
# Lookup RuleBase ID / RATING_GROUP
# bilibili
# 实际输入需要核对的RG值
lookupRuleBaseID = '1600000002'

# DPI GuiJi detail File
# 单用户的DPI明细，F列为目的IP，K列为URL
rawfile = 'MSISDN-20200407.xlsx'

# generate rule from DPI file
dpiwb = load_workbook(dpifile)
dpisheet = dpiwb.worksheets[0]
dpirownumber = dpisheet.max_row
dpicolumnnumber = dpisheet.max_column
exportflag = 1
URLlist = []
Serverlist = []
# print(dpirownumber)
# print(dpicolumnnumber)

for x in range(3, dpirownumber + 1):
    # lookup RuleBase ID column and find ip,mask,url
	# AE列为RG值，H/I列为目的IP和掩码，T列为URL
    RuleBaseID = dpisheet.cell(row=x, column=31).value
    if RuleBaseID == lookupRuleBaseID:
        ServerIP = str(dpisheet.cell(row=x, column=8).value)
        ServerIPMask = str(dpisheet.cell(row=x, column=9).value)
        URL = str(dpisheet.cell(row=x, column=20).value)
        if len(ServerIP) > 0 and len(ServerIPMask) > 0:
            Serverlist.append(ServerIP + '/' + ServerIPMask)
        if len(URL) > 0:
            URLlist.append(URL)

if len(Serverlist) > 0:
    print(lookupRuleBaseID, 'ip match rule', len(Serverlist))


if len(URLlist) > 0:
    print(lookupRuleBaseID, 'url match rule', len(URLlist))


# export RuleBase ID / RATING_GROUP to file
if exportflag == 1:
    os.chdir('config')
    exportrule(Serverlist, lookupRuleBaseID + '-ip')
    exportrule(URLlist, lookupRuleBaseID + '-url')
    os.chdir('..')

RcsIpset = netaddr.IPSet(Serverlist)
reglist = genrepatten(URLlist)

# Dealing with DPI GuiJi detail file with xGW DPI rule
inputwb = load_workbook(rawfile)
inputsheet = inputwb.worksheets[0]
rownumber = inputsheet.max_row
columnnumber = inputsheet.max_column
# print(rownumber,) columnnumber

mismatchwb = workbook.Workbook()
mismatchws = mismatchwb.worksheets[0]
matchwb = workbook.Workbook()
matchws = matchwb.worksheets[0]

copyrow(inputsheet, mismatchws, 1, 1, columnnumber)
copyrow(inputsheet, matchws, 1, 1, columnnumber)

ipmismatch = []
hostmismatch = []

bytetotal = 0
bytematch = 0
bytemismatch = 0

numbermatch = 2
numbermismatch = 2
print(u'对上网轨迹话单与PGW规则进行匹配判断', str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
for x in range(2, rownumber + 1):
    byte = inputsheet.cell(row=x, column=4).value
    serverip = inputsheet.cell(row=x, column=6).value
    url = inputsheet.cell(row=x, column=11).value
    bytetotal += byte
    # print(serverip,url)
    matchflag = 0
    # do url check
    if len(url) > 0:
        if len(reglist) > 0:
            for everyrepatten in reglist:
                if re.search(everyrepatten, url):
                    copyrow(inputsheet, matchws, x, numbermatch, columnnumber)
                    matchflag = 1
                    bytematch += byte
                    numbermatch += 1
                    break
    # url match break
    if matchflag == 1:
        continue
    else:
        # do ip rule check
        # if len(RcsIpset) > 0:
        # 20200312 IndexError: range contains more than 9223372036854775807 (sys.maxint) IP addresses
        # !Use the .size property instead.
        if RcsIpset.size > 0:
            # ip match
            if netaddr.IPAddress(serverip) in RcsIpset:
                copyrow(inputsheet, matchws, x, numbermatch, columnnumber)
                matchflag = 1
                bytematch += byte
                numbermatch += 1
                continue
            # ip mismatch
            else:
                bytemismatch += byte
                copyrow(inputsheet, mismatchws, x, numbermismatch, columnnumber)
                ipmismatch.append(netaddr.IPAddress(serverip))
                match = re.search(r"^http://([a-z0-9\-._~%]+)/", url, re.IGNORECASE)
                if match:
                    host = match.group(1)
                else:
                    host = ""
                if len(host) > 0 and host not in hostmismatch:
                    hostmismatch.append(host)
                numbermismatch += 1
                continue
        # no ip rule exist
        else:
            bytemismatch += byte
            copyrow(inputsheet, mismatchws, x, numbermismatch, columnnumber)
            ipmismatch.append(netaddr.IPAddress(serverip))
            match = re.search(r"^http://([a-z0-9\-._~%]+)/", url, re.IGNORECASE)
            if match:
                host = match.group(1)
            else:
                host = ""
            if len(host) > 0 and host not in hostmismatch:
                hostmismatch.append(host)
            numbermismatch += 1
            continue

os.chdir('output')
mismatchwb.save(str(datetime.now().strftime('%Y%m%d%H%M')) + '-' + u'未匹配' + '-' + rawfile)
matchwb.save(str(datetime.now().strftime('%Y%m%d%H%M')) + '-' + u'匹配上' + '-' + rawfile)
os.chdir('..')
print(u'上网轨迹清单文件处理完毕，输出处理后文件', str(datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
print('-' * 36)
print("%-14s %-13s %-14s " % (u'上网轨迹', u'记录数', u'流量(MB)'))
print("%-18s %-16d %-.3f " % (u'total', rownumber - 1, bytetotal / (1024 * 1024.0)))
print("%-18s %-16d %-.3f " % (u'match', numbermatch - 2, bytematch / (1024 * 1024.0)))
print("%-18s %-16d %-.3f " % (u'not match', numbermismatch - 2, bytemismatch / (1024 * 1024.0)))
print('-' * 36)
genipandhost(ipmismatch, hostmismatch)
